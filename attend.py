import cv2
import face_recognition
import openpyxl
from datetime import datetime
import os

# Faces and names
known_faces = []  # Add known face encodings
known_names = []  # Add corresponding names

image_of_person1 = face_recognition.load_image_file("person1.jpg")
person1_face_encoding = face_recognition.face_encodings(image_of_person1)[0]
known_faces.append(person1_face_encoding)
known_names.append("prakash")

image_of_person2 = face_recognition.load_image_file("person2.jpg")
person2_face_encoding = face_recognition.face_encodings(image_of_person2)[0]
known_faces.append(person2_face_encoding)
known_names.append("hari")

# Excel file process
excel_filename = 'attendance.xlsx'

# Current Date
current_date = datetime.now().strftime("%Y-%m-%d")

# Create Excel file if it doesn't exist
if not os.path.exists(excel_filename):
    workbook = openpyxl.Workbook()
    workbook.save(excel_filename)

# Load the workbook
workbook = openpyxl.load_workbook(excel_filename)

# Create a new sheet for the current date if not already present
if current_date not in workbook.sheetnames:
    sheet = workbook.create_sheet(current_date)
    sheet.append(["Name", "Roll Number", "Time", "Date"])
else:
    sheet = workbook[current_date]

# Initialize webcam
video_capture = cv2.VideoCapture(0)

while True:
    ret, frame = video_capture.read()
    
    # Convert BGR frame (OpenCV) to RGB frame (face_recognition)
    rgb_frame = frame[:, :, ::-1]
    
    # Detect faces in the current frame
    face_locations = face_recognition.face_locations(rgb_frame)
    
    # Get face encodings for each detected face
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
    
    # Process each face encoding found in the frame
    for (face_location, face_encoding) in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_faces, face_encoding)
        name = "Unknown"
        
        if True in matches:
            first_match_index = matches.index(True)
            name = known_names[first_match_index]
            
            roll_number = "roll123"  # Update with the actual roll number
            current_time = datetime.now().strftime("%H:%M:%S")
            
            # Append attendance to the sheet
            sheet.append([name, roll_number, current_time, current_date])
            workbook.save(excel_filename)
        
        # You can draw landmarks here (optional)

    # Display the video frame
    cv2.imshow('Video', frame)
    
    # Break if 'q' key is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close windows
video_capture.release()
cv2.destroyAllWindows()
