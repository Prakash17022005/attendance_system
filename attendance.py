import cv2
import face_recognition
import openpyxl
from datetime import datetime
import os

import openpyxl.workbook

#faces and names
known_faces=[]
known_names=[]

#Excel File process
excel_filename='addendance.xlsx'

#current Date
current_date=datetime.now().strftime("%Y-%m-%d")

#excel file exists or create it
if not os.path.exists(excel_filename):
    workbook=openpyxl.Workbook()
    workbook.save(excel_filename)
    
workbook=openpyxl.load_workbook(excel_filename)

if current_date not in workbook.sheetnames:
    sheet=workbook.create_sheet(current_date)
    sheet.append(["Name","Roll Number","Time","Date"])
else:
    sheet=workbook[current_date]

video_capture=cv2.VideoCapture(0)

while True:
    ret,frame=video_capture.read()
    rgb_frame=frame[:,:,::-1]
    
    face_locations=face_recognition.face_locations(rgb_frame)
    face_landmarks_list=face_recognition.face_landmarks(rgb_frame,face_locations)
    face_encodings=face_recognition.face_encodings(rgb_frame,face_locations)
    
    for face_encoding in face_encodings:
        
        matches=face_recognition.compare_faces(known_faces,face_encoding)
        
        name="unknown"
        
        if True in matches:
            first_match_index=matches.index(True)
            name=known_names[first_match_index]
            
            roll_number="roll123"
            
            current_time=datetime.now().strftime("%H:%M:%S")
            
            sheet.append([name,roll_number,current_time,current_date])
            workbook.save(excel_filename)
            
    cv2.imshow('video',frame)
    
    if cv2.waitKey(1) & 0xFF==ord('q'):
        break
    
#release the webcam
video_capture.release()
cv2.destroyAllWindows()
